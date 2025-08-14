#!/usr/bin/env python3
import json
import sys
from datetime import datetime
import openpyxl
from openpyxl.cell.cell import MergedCell
import os

def safe_set_cell_value(ws, row, col, value):
    """結合セルを考慮して安全にセルの値を設定する"""
    try:
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            # 結合セルの場合、結合範囲の左上セルを取得
            for merged_range in ws.merged_cells.ranges:
                if (merged_range.min_row <= row <= merged_range.max_row and 
                    merged_range.min_col <= col <= merged_range.max_col):
                    # 結合範囲の左上セルに値を設定
                    ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                    return True
            # 結合範囲が見つからない場合は通常のセル設定を試行
            ws.cell(row=row, column=col).value = value
        else:
            # 通常のセル
            cell.value = value
        return True
    except Exception as e:
        print(f"Warning: Could not set value for cell ({row}, {col}): {e}")
        return False

def update_excel_template(json_data):
    """JSONデータでExcelテンプレートの特定セルを更新する"""
    
    # JSONデータをパース
    data = json.loads(json_data)
    
    # Excelファイルのパス（GitHub Actions用）
    excel_path = 'data/スポーツコーダ監視作業履歴.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found: {excel_path}")
        return False
    
    try:
        # 既存のファイルを開く
        wb = openpyxl.load_workbook(excel_path)
        
        # 最初のシート（現在のシート）を使用
        ws = wb.worksheets[0]
        sheet_name = ws.title
        print(f"Updating sheet: {sheet_name}")
        
        # 最終更新日時を更新 (E2)
        safe_set_cell_value(ws, 2, 5, '最終更新日時')
        safe_set_cell_value(ws, 2, 26, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        # 基本情報を更新
        # 確認日（B5）- 時刻は含めない
        date_str = data.get('date', '')
        if date_str:
            # 日付のみを設定（時刻は除去）
            check_date = date_str.split(' ')[0]  # スペースがあれば最初の部分（日付部分）のみ
            safe_set_cell_value(ws, 5, 2, check_date)
            print(f"Updated 確認日 (B5): {check_date}")
        
        # 確認者 (C5)
        if data.get('checker'):
            safe_set_cell_value(ws, 5, 3, data.get('checker'))
            print(f"Updated 確認者 (C5): {data.get('checker')}")
        
        # 確認結果 (D5) - 入力値を優先、なければ自動判定
        check_result = data.get('checkResult')
        if check_result:
            safe_set_cell_value(ws, 5, 4, check_result)
            print(f"Updated 確認結果 (D5): {check_result}")
        else:
            # 自動判定
            has_error = False
            error_details = []
            
            # エラーチェック
            if data.get('cpuPciModule') == 'Error' or data.get('scsiEnclosure') == 'Error':
                has_error = True
                error_details.append('FTサーバーエラー')
            
            # CPU使用率チェック
            cpu_usage = data.get('cpuUsage', '')
            if cpu_usage and float(cpu_usage) > 80:
                has_error = True
                error_details.append(f'CPU高負荷({cpu_usage}%)')
            
            # 確認結果を設定
            if has_error:
                safe_set_cell_value(ws, 5, 4, '要確認')
            else:
                safe_set_cell_value(ws, 5, 4, '問題なし')
        
        # 確認内容 (E5)
        # 自動判定の場合
        has_error = False
        error_details = []
        if data.get('cpuPciModule') == 'Error' or data.get('scsiEnclosure') == 'Error':
            has_error = True
            error_details.append('FTサーバーエラー')
        cpu_usage = data.get('cpuUsage', '')
        if cpu_usage and float(cpu_usage) > 80:
            has_error = True
            error_details.append(f'CPU高負荷({cpu_usage}%)')
        
        if error_details:
            safe_set_cell_value(ws, 5, 5, '; '.join(error_details))
        else:
            safe_set_cell_value(ws, 5, 5, 'FTサーバユーティリティ確認')
        
        # FTサーバユーティリティ
        # CPU/PCIモジュール (P6)
        if data.get('cpuPciModule'):
            safe_set_cell_value(ws, 6, 16, data.get('cpuPciModule'))
        
        # SCSIエンクロージャ (P7)
        if data.get('scsiEnclosure'):
            safe_set_cell_value(ws, 7, 16, data.get('scsiEnclosure'))
        
        # HDD残容量
        # Cドライブ (G11)
        if data.get('driveC'):
            safe_set_cell_value(ws, 11, 7, float(data.get('driveC')))
        
        # Dドライブ (Q11)
        if data.get('driveD'):
            safe_set_cell_value(ws, 11, 17, float(data.get('driveD')))
        
        # Eドライブ (G12)
        if data.get('driveE'):
            safe_set_cell_value(ws, 12, 7, float(data.get('driveE')))
        
        # Yドライブ (Q12)
        if data.get('driveY'):
            safe_set_cell_value(ws, 12, 17, float(data.get('driveY')))
        
        # Zドライブ (G13)
        if data.get('driveZ'):
            safe_set_cell_value(ws, 13, 7, float(data.get('driveZ')))
        
        # メモリ・CPU使用状況
        # SQLServerメモリ (P15)
        if data.get('sqlServerMemory'):
            safe_set_cell_value(ws, 15, 16, float(data.get('sqlServerMemory')))
        
        # 全体メモリ使用量 (P16)
        if data.get('totalMemory'):
            safe_set_cell_value(ws, 16, 16, float(data.get('totalMemory')))
        
        # CPU使用率 (J18)
        if data.get('cpuUsage'):
            safe_set_cell_value(ws, 18, 10, float(data.get('cpuUsage')))
        
        # CPU確認時刻 (P18)
        if data.get('cpuUsageTime'):
            safe_set_cell_value(ws, 18, 16, data.get('cpuUsageTime'))
        
        # サーバ時刻同期状況 (適切な場所に配置) 
        if data.get('serverTimeSync'):
            # J9セルの「問題なし」欄に配置
            safe_set_cell_value(ws, 9, 10, data.get('serverTimeSync'))
            print(f"Updated サーバ時刻同期 (J9): {data.get('serverTimeSync')}")
        
        # メモリ使用量状況 (適切な場所に配置)
        if data.get('memoryUsageStatus'):
            # J17セルの「問題なし」欄に配置
            safe_set_cell_value(ws, 17, 10, data.get('memoryUsageStatus'))
            print(f"Updated メモリ使用量状況 (J17): {data.get('memoryUsageStatus')}")
        
        # サーバーランプ
        # 上段サーバランプ (M20)
        if data.get('upperLamps'):
            safe_set_cell_value(ws, 20, 13, int(data.get('upperLamps')))
            print(f"Updated 上段サーバーランプ (M20): {data.get('upperLamps')}個")
        
        # 下段サーバーランプ - M21を試行（要確認）
        if data.get('lowerLamps'):
            safe_set_cell_value(ws, 21, 13, int(data.get('lowerLamps')))
            print(f"Updated 下段サーバーランプ (M21): {data.get('lowerLamps')}個")
        
        # 備考欄をZ列に更新
        if data.get('notes'):
            # Z列（26列目）の適切な行に備考を設定
            # 基本情報の行（5行目）に合わせてZ5に設定
            safe_set_cell_value(ws, 5, 26, data.get('notes'))
            print(f"Updated 備考 (Z5): {data.get('notes')}")
        
        # SC機のHDD容量を更新
        # SC機のセル位置を特定して更新する
        update_sc_machines(ws, data)
        
        # ファイルを保存
        wb.save(excel_path)
        print(f"✓ Excel template updated successfully: {excel_path}")
        print(f"Updated cells: 基本情報, FTサーバー, HDD容量, メモリ・CPU, サーバーランプ")
        
        return True
        
    except Exception as e:
        print(f"Error updating Excel template: {e}")
        return False

def update_sc_machines(ws, data):
    """SC機のHDD容量を更新する"""
    
    # SC機のデータの開始位置: E24からE33 (SC-1からSC-10)
    start_row = 24
    
    for i in range(1, 13):  # SC-1 から SC-12まで拡張
        # SC機のC, D, Hドライブデータがあるかチェック
        sc_c = data.get(f'sc{i}_c')
        sc_d = data.get(f'sc{i}_d')
        sc_h = data.get(f'sc{i}_h')
        
        if sc_c or sc_d or sc_h:
            # SC機の行位置を計算: E24=SC-1, E25=SC-2, ...
            sc_row = start_row + (i - 1)
            
            # SC機のラベルを確認/設定 (E列)
            current_label = ws.cell(row=sc_row, column=5).value
            if not current_label or f'SC-{i}' not in str(current_label):
                safe_set_cell_value(ws, sc_row, 5, f'SC-{i}')
            
            # HDD容量を別々の列に配置
            # HIJ列: Cドライブ
            if sc_c:
                safe_set_cell_value(ws, sc_row, 8, float(sc_c))  # H列: C値
                print(f"Updated SC-{i} C-drive (H{sc_row}): {sc_c}GB")
            
            # MNO列: Dドライブ
            if sc_d:
                safe_set_cell_value(ws, sc_row, 13, float(sc_d))  # M列: D値
                print(f"Updated SC-{i} D-drive (M{sc_row}): {sc_d}GB")
            
            # RST列: Hドライブ
            if sc_h:
                safe_set_cell_value(ws, sc_row, 18, float(sc_h))  # R列: H値
                print(f"Updated SC-{i} H-drive (R{sc_row}): {sc_h}GB")
    
    # CPU[%]/メモリ[GB]セクションの更新
    update_sc_cpu_memory(ws, data)

def update_sc_cpu_memory(ws, data):
    """SC機のCPU[%]/メモリ[GB]を更新する"""
    
    # CPU[%]/メモリ[GB]セクションの開始位置を探す
    cpu_section_row = None
    for row in range(35, 50):
        cell_value = ws.cell(row=row, column=5).value
        if cell_value and 'CPU[%]/メモリ[GB]' in str(cell_value):
            cpu_section_row = row
            break
    
    if not cpu_section_row:
        print("CPU[%]/メモリ[GB] section not found")
        return
    
    print(f"Found CPU[%]/メモリ[GB] section at row {cpu_section_row}")
    
    # SC-1からSC-12まで処理
    for i in range(1, 13):
        sc_cpu = data.get(f'sc{i}_cpu')
        sc_memory = data.get(f'sc{i}_memory')
        
        if sc_cpu or sc_memory:
            # 奇数番号のSC (SC-1, SC-3, SC-5, ..., SC-11) は左側の列
            # 偶数番号のSC (SC-2, SC-4, SC-6, ..., SC-12) は右側の列
            if i % 2 == 1:  # 奇数
                # 左側の列 (E, G, I)
                row_offset = (i - 1) // 2
                target_row = cpu_section_row + 1 + row_offset
                
                # SC-X: ラベル (E列)
                safe_set_cell_value(ws, target_row, 5, f'SC-{i}:')
                
                # CPU使用率 (G列)
                if sc_cpu:
                    safe_set_cell_value(ws, target_row, 7, float(sc_cpu))
                    print(f"Updated SC-{i} CPU% (G{target_row}): {sc_cpu}")
                
                # メモリ使用量 (K列) - KL列の範囲なのでK列を使用
                if sc_memory:
                    safe_set_cell_value(ws, target_row, 11, float(sc_memory))
                    print(f"Updated SC-{i} Memory (K{target_row}): {sc_memory}GB")
                
            else:  # 偶数
                # 右側の列 (O, Q, U)
                row_offset = (i - 2) // 2
                target_row = cpu_section_row + 1 + row_offset
                
                # SC-X: ラベル (O列)
                safe_set_cell_value(ws, target_row, 15, f'SC-{i}:')
                
                # CPU使用率 (Q列)
                if sc_cpu:
                    safe_set_cell_value(ws, target_row, 17, float(sc_cpu))
                    print(f"Updated SC-{i} CPU% (Q{target_row}): {sc_cpu}")
                
                # メモリ使用量 (U列) - UV列の範囲なのでU列を使用
                if sc_memory:
                    safe_set_cell_value(ws, target_row, 21, float(sc_memory))
                    print(f"Updated SC-{i} Memory (U{target_row}): {sc_memory}GB")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python update_excel_template.py '<json_data>' or python update_excel_template.py <json_file_path>")
        sys.exit(1)
    
    json_input = sys.argv[1]
    
    # ファイルパスかJSONデータかを判定
    if os.path.exists(json_input):
        # ファイルパスの場合、ファイルを読み込み
        print(f"Reading JSON file: {json_input}")
        with open(json_input, 'r', encoding='utf-8') as f:
            json_data = f.read()
    else:
        # JSON文字列の場合
        json_data = json_input
    
    success = update_excel_template(json_data)
    if not success:
        sys.exit(1)